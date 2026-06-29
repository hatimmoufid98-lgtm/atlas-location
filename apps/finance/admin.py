from django.contrib import admin
from django.utils.html import format_html
from django.db.models import Sum
from django.http import HttpResponse
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from .models import CategorieDepense, Depense


def export_depenses_excel(modeladmin, request, queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Depenses"

    header_fill = PatternFill("solid", fgColor="C8A96E")
    header_font = Font(bold=True, color="FFFFFF")
    headers = ["Date", "Catégorie", "Description", "Montant (MAD)", "Agence", "Notes"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    total = 0
    for row, d in enumerate(queryset, 2):
        ws.cell(row=row, column=1, value=str(d.date))
        ws.cell(row=row, column=2, value=str(d.categorie))
        ws.cell(row=row, column=3, value=d.description)
        ws.cell(row=row, column=4, value=float(d.montant))
        ws.cell(row=row, column=5, value=str(d.agency))
        ws.cell(row=row, column=6, value=d.notes)
        total += float(d.montant)

    last_row = queryset.count() + 2
    ws.cell(row=last_row, column=3, value="TOTAL").font = Font(bold=True)
    total_cell = ws.cell(row=last_row, column=4, value=round(total, 2))
    total_cell.font = Font(bold=True)
    total_cell.fill = PatternFill("solid", fgColor="FFF3CD")

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="depenses.xlsx"'
    wb.save(response)
    return response

export_depenses_excel.short_description = "Exporter en Excel"


@admin.register(CategorieDepense)
class CategorieDepenseAdmin(admin.ModelAdmin):
    list_display = ('nom', 'icone', 'description')
    search_fields = ('nom',)


@admin.register(Depense)
class DepenseAdmin(admin.ModelAdmin):
    list_display = ('date', 'categorie', 'description', 'montant_display', 'agency', 'a_justificatif')
    list_filter = ('agency', 'categorie', 'date')
    search_fields = ('description', 'notes')
    date_hierarchy = 'date'
    actions = [export_depenses_excel]

    fieldsets = (
        ('Dépense', {'fields': ('agency', 'date', 'categorie', 'description', 'montant')}),
        ('Documents', {'fields': ('justificatif', 'notes')}),
    )

    def montant_display(self, obj):
        return format_html('<strong>{} MAD</strong>', obj.montant)
    montant_display.short_description = "Montant"
    montant_display.admin_order_field = 'montant'

    def a_justificatif(self, obj):
        if obj.justificatif:
            return format_html('<span style="color:#10b981;">Oui</span>')
        return format_html('<span style="color:#9ca3af;">Non</span>')
    a_justificatif.short_description = "Justificatif"

    def get_queryset(self, request):
        return super().get_queryset(request).select_related('agency', 'categorie')
