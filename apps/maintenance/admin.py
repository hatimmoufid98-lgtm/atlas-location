from django.contrib import admin
from django.utils.html import format_html
from django.utils import timezone
from django.http import HttpResponse
from django.db.models import Sum
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from .models import Assurance, Vignette, Lavage, PleinCarburant, Revision


def badge_alerte(jours):
    if jours < 0:
        return format_html('<span style="background:#ef4444;color:white;padding:2px 8px;border-radius:12px;font-size:11px;">Expiré</span>')
    elif jours <= 30:
        return format_html('<span style="background:#f59e0b;color:white;padding:2px 8px;border-radius:12px;font-size:11px;">⚠ {} jours</span>', jours)
    else:
        return format_html('<span style="background:#10b981;color:white;padding:2px 8px;border-radius:12px;font-size:11px;">✓ {} jours</span>', jours)


# ─── ASSURANCE ───────────────────────────────────────────────────────────────

def export_assurances_excel(modeladmin, request, queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Assurances"

    header_fill = PatternFill("solid", fgColor="C8A96E")
    header_font = Font(bold=True, color="FFFFFF")
    headers = ["Voiture", "Compagnie", "N° Police", "Début", "Expiration", "Montant (MAD)", "Jours restants", "Statut"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row, a in enumerate(queryset, 2):
        ws.cell(row=row, column=1, value=str(a.car))
        ws.cell(row=row, column=2, value=a.compagnie)
        ws.cell(row=row, column=3, value=a.numero_police)
        ws.cell(row=row, column=4, value=str(a.date_debut))
        ws.cell(row=row, column=5, value=str(a.date_fin))
        ws.cell(row=row, column=6, value=float(a.montant_annuel))
        ws.cell(row=row, column=7, value=a.jours_restants)
        ws.cell(row=row, column=8, value="Active" if a.est_active else "Expirée")

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="assurances.xlsx"'
    wb.save(response)
    return response

export_assurances_excel.short_description = "Exporter en Excel"


@admin.register(Assurance)
class AssuranceAdmin(admin.ModelAdmin):
    list_display = ('car', 'compagnie', 'numero_police', 'date_debut', 'date_fin', 'montant_annuel', 'statut_badge')
    list_filter = ('car__agency', 'compagnie')
    search_fields = ('car__marque', 'car__modele', 'compagnie', 'numero_police')
    date_hierarchy = 'date_fin'
    actions = [export_assurances_excel]

    fieldsets = (
        ('Véhicule', {'fields': ('car',)}),
        ("Détails de l'assurance", {'fields': ('compagnie', 'numero_police', 'date_debut', 'date_fin', 'montant_annuel')}),
        ('Documents', {'fields': ('document', 'notes')}),
    )

    def statut_badge(self, obj):
        return badge_alerte(obj.jours_restants)
    statut_badge.short_description = "Statut"

    def get_queryset(self, request):
        return super().get_queryset(request).select_related('car', 'car__agency')


# ─── VIGNETTE ────────────────────────────────────────────────────────────────

def export_vignettes_excel(modeladmin, request, queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vignettes"

    header_fill = PatternFill("solid", fgColor="C8A96E")
    header_font = Font(bold=True, color="FFFFFF")
    headers = ["Voiture", "Année", "Date paiement", "Date expiration", "Montant (MAD)", "Jours restants"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    for row, v in enumerate(queryset, 2):
        ws.cell(row=row, column=1, value=str(v.car))
        ws.cell(row=row, column=2, value=v.annee)
        ws.cell(row=row, column=3, value=str(v.date_paiement))
        ws.cell(row=row, column=4, value=str(v.date_expiration))
        ws.cell(row=row, column=5, value=float(v.montant))
        ws.cell(row=row, column=6, value=v.jours_restants)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="vignettes.xlsx"'
    wb.save(response)
    return response

export_vignettes_excel.short_description = "Exporter en Excel"


@admin.register(Vignette)
class VignetteAdmin(admin.ModelAdmin):
    list_display = ('car', 'annee', 'date_paiement', 'date_expiration', 'montant', 'statut_badge')
    list_filter = ('car__agency', 'annee')
    search_fields = ('car__marque', 'car__modele')
    actions = [export_vignettes_excel]

    def statut_badge(self, obj):
        return badge_alerte(obj.jours_restants)
    statut_badge.short_description = "Statut"


# ─── LAVAGE ──────────────────────────────────────────────────────────────────

def export_lavages_excel(modeladmin, request, queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lavages"

    header_fill = PatternFill("solid", fgColor="C8A96E")
    header_font = Font(bold=True, color="FFFFFF")
    headers = ["Voiture", "Date", "Type", "Montant (MAD)", "Prestataire"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    total = 0
    for row, l in enumerate(queryset, 2):
        ws.cell(row=row, column=1, value=str(l.car))
        ws.cell(row=row, column=2, value=str(l.date))
        ws.cell(row=row, column=3, value=l.get_type_lavage_display())
        ws.cell(row=row, column=4, value=float(l.montant))
        ws.cell(row=row, column=5, value=l.prestataire)
        total += float(l.montant)

    last_row = queryset.count() + 2
    ws.cell(row=last_row, column=3, value="TOTAL").font = Font(bold=True)
    ws.cell(row=last_row, column=4, value=total).font = Font(bold=True)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="lavages.xlsx"'
    wb.save(response)
    return response

export_lavages_excel.short_description = "Exporter en Excel"


@admin.register(Lavage)
class LavageAdmin(admin.ModelAdmin):
    list_display = ('car', 'date', 'type_lavage', 'montant', 'prestataire')
    list_filter = ('car__agency', 'type_lavage', 'date')
    search_fields = ('car__marque', 'car__modele', 'prestataire')
    date_hierarchy = 'date'
    actions = [export_lavages_excel]


# ─── PLEIN CARBURANT ─────────────────────────────────────────────────────────

def export_carburant_excel(modeladmin, request, queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Carburant"

    header_fill = PatternFill("solid", fgColor="C8A96E")
    header_font = Font(bold=True, color="FFFFFF")
    headers = ["Voiture", "Date", "Km compteur", "Litres", "Montant (MAD)", "Carburant", "Station", "Conso (L/100km)"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    total_montant = 0
    total_litres = 0
    for row, p in enumerate(queryset, 2):
        conso = p.consommation_aux_100
        ws.cell(row=row, column=1, value=str(p.car))
        ws.cell(row=row, column=2, value=str(p.date))
        ws.cell(row=row, column=3, value=p.km_compteur)
        ws.cell(row=row, column=4, value=float(p.litres))
        ws.cell(row=row, column=5, value=float(p.montant))
        ws.cell(row=row, column=6, value=p.get_type_carburant_display())
        ws.cell(row=row, column=7, value=p.station)
        ws.cell(row=row, column=8, value=conso if conso else "—")
        total_montant += float(p.montant)
        total_litres += float(p.litres)

    last_row = queryset.count() + 2
    ws.cell(row=last_row, column=3, value="TOTAL").font = Font(bold=True)
    ws.cell(row=last_row, column=4, value=round(total_litres, 2)).font = Font(bold=True)
    ws.cell(row=last_row, column=5, value=round(total_montant, 2)).font = Font(bold=True)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="carburant.xlsx"'
    wb.save(response)
    return response

export_carburant_excel.short_description = "Exporter en Excel"


@admin.register(PleinCarburant)
class PleinCarburantAdmin(admin.ModelAdmin):
    list_display = ('car', 'date', 'km_compteur', 'litres', 'montant', 'type_carburant', 'conso_display', 'station')
    list_filter = ('car__agency', 'type_carburant', 'date')
    search_fields = ('car__marque', 'car__modele', 'station')
    date_hierarchy = 'date'
    actions = [export_carburant_excel]

    def conso_display(self, obj):
        conso = obj.consommation_aux_100
        if conso:
            return f"{conso} L/100km"
        return "—"
    conso_display.short_description = "Consommation"


# ─── RÉVISION ────────────────────────────────────────────────────────────────

def export_revisions_excel(modeladmin, request, queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Revisions"

    header_fill = PatternFill("solid", fgColor="C8A96E")
    header_font = Font(bold=True, color="FFFFFF")
    headers = ["Voiture", "Date", "Type", "Km", "Montant (MAD)", "Garage", "Prochaine révision (km)", "Prochaine révision (date)"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    total = 0
    for row, r in enumerate(queryset, 2):
        ws.cell(row=row, column=1, value=str(r.car))
        ws.cell(row=row, column=2, value=str(r.date))
        ws.cell(row=row, column=3, value=r.get_type_revision_display())
        ws.cell(row=row, column=4, value=r.km_compteur)
        ws.cell(row=row, column=5, value=float(r.montant))
        ws.cell(row=row, column=6, value=r.garage)
        ws.cell(row=row, column=7, value=r.prochaine_revision_km or "—")
        ws.cell(row=row, column=8, value=str(r.prochaine_revision_date) if r.prochaine_revision_date else "—")
        total += float(r.montant)

    last_row = queryset.count() + 2
    ws.cell(row=last_row, column=4, value="TOTAL").font = Font(bold=True)
    ws.cell(row=last_row, column=5, value=round(total, 2)).font = Font(bold=True)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="revisions.xlsx"'
    wb.save(response)
    return response

export_revisions_excel.short_description = "Exporter en Excel"


@admin.register(Revision)
class RevisionAdmin(admin.ModelAdmin):
    list_display = ('car', 'date', 'type_revision', 'km_compteur', 'montant', 'garage', 'prochaine_revision_km', 'prochaine_revision_date')
    list_filter = ('car__agency', 'type_revision', 'date')
    search_fields = ('car__marque', 'car__modele', 'garage')
    date_hierarchy = 'date'
    actions = [export_revisions_excel]

    fieldsets = (
        ('Véhicule', {'fields': ('car',)}),
        ('Intervention', {'fields': ('date', 'type_revision', 'description', 'km_compteur', 'montant', 'garage')}),
        ('Prochaine révision', {'fields': ('prochaine_revision_km', 'prochaine_revision_date')}),
        ('Document', {'fields': ('facture',)}),
    )
