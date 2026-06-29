from datetime import date
from decimal import Decimal

from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render
from django.db.models import Sum
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from apps.agencies.models import Agency
from apps.cars.models import Car
from apps.bookings.models import Booking, BlockedPeriod
from apps.maintenance.models import Assurance, Vignette, Lavage, PleinCarburant, Revision
from apps.finance.models import Depense

MOIS_FR = [
    "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
    "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre",
]

MOIS_FR_COURT = ["Jan", "Fév", "Mar", "Avr", "Mai", "Juin", "Juil", "Aoû", "Sep", "Oct", "Nov", "Déc"]


def _recettes_par_mois(annee):
    """Recettes mensuelles des reservations confirmees, attribuees au mois de debut."""
    recettes = [Decimal('0')] * 12
    bookings = Booking.objects.filter(
        statut='confirmee', date_debut__year=annee
    ).select_related('car')
    for b in bookings:
        recettes[b.date_debut.month - 1] += Decimal(str(b.prix_total))
    return recettes


def _depenses_agence_par_mois(annee):
    depenses = [Decimal('0')] * 12
    qs = Depense.objects.filter(date__year=annee)
    for d in qs:
        depenses[d.date.month - 1] += d.montant
    return depenses


def _charges_vehicules_par_mois(annee):
    """Lavages + carburant + revisions par mois (depenses reelles datees)."""
    charges = [Decimal('0')] * 12
    for model in (Lavage, PleinCarburant, Revision):
        for obj in model.objects.filter(date__year=annee):
            charges[obj.date.month - 1] += obj.montant
    return charges


@staff_member_required
def tableau_bord(request):
    today = date.today()
    annee = int(request.GET.get('annee', today.year))

    # ─── ALERTES ──────────────────────────────────────────────────────────────
    assurances = Assurance.objects.select_related('car').all()
    assurances_alertes = sorted(
        [a for a in assurances if a.jours_restants <= 30],
        key=lambda a: a.jours_restants,
    )
    vignettes = Vignette.objects.select_related('car').all()
    vignettes_alertes = sorted(
        [v for v in vignettes if v.jours_restants <= 30],
        key=lambda v: v.jours_restants,
    )
    # Revisions dont la prochaine date est dans <= 30 jours (ou depassee)
    from datetime import timedelta
    limite = today + timedelta(days=30)
    revisions_a_venir = list(
        Revision.objects.select_related('car')
        .filter(prochaine_revision_date__isnull=False, prochaine_revision_date__lte=limite)
        .order_by('prochaine_revision_date')
    )

    # ─── BILAN ANNUEL ─────────────────────────────────────────────────────────
    recettes_mois = _recettes_par_mois(annee)
    depenses_agence_mois = _depenses_agence_par_mois(annee)
    charges_veh_mois = _charges_vehicules_par_mois(annee)

    bilan_mensuel = []
    for i in range(12):
        depenses_totales = depenses_agence_mois[i] + charges_veh_mois[i]
        benefice = recettes_mois[i] - depenses_totales
        bilan_mensuel.append({
            'mois': MOIS_FR[i],
            'recettes': recettes_mois[i],
            'depenses': depenses_totales,
            'benefice': benefice,
        })

    total_recettes = sum(recettes_mois)
    total_depenses_agence = sum(depenses_agence_mois)
    total_charges_veh = sum(charges_veh_mois)
    total_depenses = total_depenses_agence + total_charges_veh
    benefice_net = total_recettes - total_depenses

    # ─── REPARTITION DES CHARGES ──────────────────────────────────────────────
    depenses_par_categorie = list(
        Depense.objects.filter(date__year=annee)
        .values('categorie__nom', 'categorie__icone')
        .annotate(total=Sum('montant'))
        .order_by('-total')
    )
    charges_vehicules_detail = [
        {'nom': 'Assurances', 'total': Assurance.objects.aggregate(s=Sum('montant_annuel'))['s'] or Decimal('0')},
        {'nom': 'Vignettes', 'total': Vignette.objects.filter(annee=annee).aggregate(s=Sum('montant'))['s'] or Decimal('0')},
        {'nom': 'Lavages', 'total': Lavage.objects.filter(date__year=annee).aggregate(s=Sum('montant'))['s'] or Decimal('0')},
        {'nom': 'Carburant', 'total': PleinCarburant.objects.filter(date__year=annee).aggregate(s=Sum('montant'))['s'] or Decimal('0')},
        {'nom': 'Révisions', 'total': Revision.objects.filter(date__year=annee).aggregate(s=Sum('montant'))['s'] or Decimal('0')},
    ]

    # Annees disponibles pour le selecteur
    annees = sorted(set(
        list(Booking.objects.dates('date_debut', 'year').values_list('date_debut__year', flat=True)) +
        [today.year, today.year - 1]
    ), reverse=True)

    nb_reservations = Booking.objects.filter(statut='confirmee', date_debut__year=annee).count()

    context = {
        'annee': annee,
        'annees': annees,
        'today': today,
        'nb_voitures': Car.objects.count(),
        'nb_reservations': nb_reservations,
        # alertes
        'assurances_alertes': assurances_alertes,
        'vignettes_alertes': vignettes_alertes,
        'revisions_a_venir': revisions_a_venir,
        'nb_alertes': len(assurances_alertes) + len(vignettes_alertes) + len(revisions_a_venir),
        # bilan
        'bilan_mensuel': bilan_mensuel,
        'total_recettes': total_recettes,
        'total_depenses': total_depenses,
        'total_depenses_agence': total_depenses_agence,
        'total_charges_veh': total_charges_veh,
        'benefice_net': benefice_net,
        # repartition
        'depenses_par_categorie': depenses_par_categorie,
        'charges_vehicules_detail': charges_vehicules_detail,
    }
    return render(request, 'dashboard/tableau_bord.html', context)


@staff_member_required
def export_bilan_excel(request):
    today = date.today()
    annee = int(request.GET.get('annee', today.year))

    recettes_mois = _recettes_par_mois(annee)
    depenses_agence_mois = _depenses_agence_par_mois(annee)
    charges_veh_mois = _charges_vehicules_par_mois(annee)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Bilan {annee}"

    header_fill = PatternFill("solid", fgColor="C8A96E")
    header_font = Font(bold=True, color="FFFFFF")
    titre_font = Font(bold=True, size=14)

    ws.cell(row=1, column=1, value=f"BILAN ANNUEL {annee} — Atlas Location").font = titre_font

    headers = ["Mois", "Recettes (MAD)", "Dépenses (MAD)", "Bénéfice net (MAD)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    total_r = total_d = Decimal('0')
    for i in range(12):
        depenses = depenses_agence_mois[i] + charges_veh_mois[i]
        benefice = recettes_mois[i] - depenses
        row = 4 + i
        ws.cell(row=row, column=1, value=MOIS_FR[i])
        ws.cell(row=row, column=2, value=float(recettes_mois[i]))
        ws.cell(row=row, column=3, value=float(depenses))
        bcell = ws.cell(row=row, column=4, value=float(benefice))
        if benefice < 0:
            bcell.font = Font(color="C00000")
        total_r += recettes_mois[i]
        total_d += depenses

    total_row = 16
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=2, value=float(total_r)).font = Font(bold=True)
    ws.cell(row=total_row, column=3, value=float(total_d)).font = Font(bold=True)
    bnet = ws.cell(row=total_row, column=4, value=float(total_r - total_d))
    bnet.font = Font(bold=True, size=12)
    bnet.fill = PatternFill("solid", fgColor="D4EDDA" if total_r - total_d >= 0 else "F8D7DA")

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="bilan_{annee}.xlsx"'
    wb.save(response)
    return response


# ─── PLANNING ANNUEL (vue Gantt) ─────────────────────────────────────────────

def _segment_pct(d_start, d_end, an_start, total_jours):
    """Retourne (left%, width%) d'un segment borne a l'annee, ou None si hors annee."""
    from datetime import date as _date
    an_end = _date(an_start.year, 12, 31)
    s = max(d_start, an_start)
    e = min(d_end, an_end)
    if e < s:
        return None
    left = (s - an_start).days / total_jours * 100
    width = ((e - s).days + 1) / total_jours * 100
    return round(left, 3), round(max(width, 0.3), 3)


@staff_member_required
def planning_annuel(request):
    today = date.today()
    annee = int(request.GET.get('annee', today.year))

    an_start = date(annee, 1, 1)
    an_end = date(annee, 12, 31)
    total_jours = (an_end - an_start).days + 1

    cars = list(Car.objects.select_related('agency').all())

    # Reservations actives qui chevauchent l'annee
    bookings = Booking.objects.filter(
        statut__in=['confirmee', 'en_attente'],
        date_debut__lte=an_end, date_fin__gte=an_start,
    ).select_related('car')
    blocked = BlockedPeriod.objects.filter(
        date_debut__lte=an_end, date_fin__gte=an_start,
    ).select_related('car')

    # Regroupement par voiture
    segments_par_car = {c.id: [] for c in cars}
    jours_loues = {c.id: 0 for c in cars}
    recettes_car = {c.id: Decimal('0') for c in cars}

    for b in bookings:
        seg = _segment_pct(b.date_debut, b.date_fin, an_start, total_jours)
        if not seg:
            continue
        left, width = seg
        if b.statut == 'confirmee':
            cls, couleur = 'confirmee', '#10b981'
            # compteurs uniquement pour confirmees
            s = max(b.date_debut, an_start)
            e = min(b.date_fin, an_end)
            jours_loues[b.car_id] += (e - s).days
            recettes_car[b.car_id] += Decimal(str(b.prix_total))
        else:
            cls, couleur = 'attente', '#f59e0b'
        segments_par_car[b.car_id].append({
            'left': left, 'width': width, 'cls': cls, 'couleur': couleur,
            'tooltip': f"{b.nom_client} — {b.date_debut:%d/%m} au {b.date_fin:%d/%m} ({b.get_statut_display()})",
        })

    for bp in blocked:
        seg = _segment_pct(bp.date_debut, bp.date_fin, an_start, total_jours)
        if not seg:
            continue
        left, width = seg
        segments_par_car[bp.car_id].append({
            'left': left, 'width': width, 'cls': 'bloque', 'couleur': '#ef4444',
            'tooltip': f"Bloquée — {bp.raison} ({bp.date_debut:%d/%m} au {bp.date_fin:%d/%m})",
        })

    jours_ecoules = (min(today, an_end) - an_start).days + 1 if today >= an_start else 0

    planning = []
    for c in cars:
        taux = round(jours_loues[c.id] / jours_ecoules * 100) if jours_ecoules > 0 else 0
        planning.append({
            'car': c,
            'segments': segments_par_car[c.id],
            'taux_occupation': taux,
            'recettes': recettes_car[c.id],
            'nb_segments': len(segments_par_car[c.id]),
        })

    # Position de la ligne "aujourd'hui"
    today_pct = None
    if an_start <= today <= an_end:
        today_pct = round((today - an_start).days / total_jours * 100, 3)

    annees = sorted(set(
        list(Booking.objects.dates('date_debut', 'year').values_list('date_debut__year', flat=True)) +
        [today.year, today.year - 1, today.year + 1]
    ), reverse=True)

    context = {
        'annee': annee,
        'annees': annees,
        'today': today,
        'today_pct': today_pct,
        'mois_courts': MOIS_FR_COURT,
        'planning': planning,
        'nb_voitures': len(cars),
    }
    return render(request, 'dashboard/planning_annuel.html', context)
